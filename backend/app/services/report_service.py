from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.ticket import Ticket


def get_all_tickets(db: Session):
    return db.query(Ticket).all()


def export_tickets_to_excel(db: Session):
    tickets = db.query(Ticket).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tickets"

    headers = [
        "ID",
        "Title",
        "Description",
        "Department",
        "Priority",
        "Status",
        "Assigned To",
        "Employee",
        "Created At"
    ]

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Ticket data
    for row, ticket in enumerate(tickets, start=2):
        sheet.cell(row=row, column=1).value = ticket.id
        sheet.cell(row=row, column=2).value = ticket.title
        sheet.cell(row=row, column=3).value = ticket.description
        sheet.cell(row=row, column=4).value = ticket.department
        sheet.cell(row=row, column=5).value = ticket.priority
        sheet.cell(row=row, column=6).value = ticket.status
        sheet.cell(row=row, column=7).value = ticket.assigned_to
        sheet.cell(row=row, column=8).value = ticket.employee_name
        sheet.cell(row=row, column=9).value = (
            ticket.created_at.strftime("%Y-%m-%d %H:%M")
            if ticket.created_at else ""
        )

    file_name = "tickets.xlsx"
    workbook.save(file_name)

    return file_name